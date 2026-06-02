"""
output/excel_exporter.py
Generates an .xlsx file in memory and returns raw bytes.
No storage service needed — the backend streams it directly to the browser.
"""
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PRIORITY_COLORS = {
    "Hot":    "FFDD4444",
    "Warm":   "FFFF8C00",
    "Medium": "FFFFD700",
    "Cold":   "FF90EE90",
    "Skip":   "FFD3D3D3",
}

COLUMNS = [
    ("Business Name",    "name",              30),
    ("Category",         "query",             20),
    ("City",             "city",              14),
    ("Area",             "area",              18),
    ("Phone",            "phone",             16),
    ("Address",          "address",           35),
    ("Lead Type",        "lead_type",         28),
    ("Website",          "website",           30),
    ("Website Domain",   "website_domain",    22),
    ("Website Status",   "website_status",    14),
    ("Instagram",        "instagram_handle",  20),
    ("Confidence",       "confidence",        12),
    ("Evidence",         "evidence",          45),
    ("Source",           "source",            16),
    ("Source Query",     "source_query",      22),
    ("Source Area",      "source_area",       18),
    ("Google Maps",      "google_maps_url",   30),
    ("Score",            "score",              8),
    ("Priority",         "priority",          10),
]

PRIORITY_ORDER = {"Hot": 0, "Warm": 1, "Medium": 2, "Cold": 3, "Skip": 4}


def build_xlsx(leads: list[dict]) -> bytes:
    """
    Build an .xlsx file from a list of lead dicts.
    Returns raw bytes — caller streams these to the browser.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header
    header_fill = PatternFill("solid", fgColor="FF111827")
    header_font = Font(color="FFFFFFFF", bold=True, name="Calibri", size=10)
    for col_idx, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    # Sort: Hot → Warm → Medium → Cold → Skip, then by score desc
    sorted_leads = sorted(
        leads,
        key=lambda l: (PRIORITY_ORDER.get(l.get("priority", "Skip"), 4), -(l.get("score") or 0))
    )

    # Rows
    row_font = Font(name="Calibri", size=9)
    for row_idx, lead in enumerate(sorted_leads, 2):
        priority = lead.get("priority", "Skip")
        row_fill = PatternFill("solid", fgColor=PRIORITY_COLORS.get(priority, "FFFFFFFF"))
        for col_idx, (_, field, _) in enumerate(COLUMNS, 1):
            value = lead.get(field) or ""
            if field == "instagram_handle" and value:
                value = f"https://instagram.com/{value}"
                
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")
            
            if value and field in ["website", "instagram_handle", "google_maps_url"]:
                cell.hyperlink = value
                cell.font = Font(color="0000FF", underline="single", size=9)
            else:
                cell.font = row_font

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
